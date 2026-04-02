import csv
import re
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook


def safe_name(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', '_', name).strip() or 'sheet'


def ask_xlsx_path() -> Path:
    while True:
        raw = input("请输入xlsx文件路径（可直接拖入终端）: ").strip().strip('"')
        path = Path(raw)

        if not path.exists():
            print("文件不存在，请重新输入。")
            continue

        if path.suffix.lower() != ".xlsx":
            print("仅支持 .xlsx 文件，请重新输入。")
            continue

        return path


def ask_output_dir(default_dir: Path) -> Path:
    raw = input(f"请输入CSV输出目录（直接回车默认: {default_dir}）: ").strip().strip('"')
    if not raw:
        return default_dir

    output_dir = Path(raw)
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def ask_sheet_mode(sheet_names: List[str]) -> Optional[str]:
    if len(sheet_names) == 1:
        print(f"检测到单工作表: {sheet_names[0]}")
        return sheet_names[0]

    print("\n检测到多个工作表:")
    for i, name in enumerate(sheet_names, start=1):
        print(f"  {i}. {name}")

    while True:
        choice = input("请选择转换方式：1=全部工作表, 2=单个工作表: ").strip()
        if choice == "1":
            return None
        if choice == "2":
            idx = input("请输入工作表序号: ").strip()
            if idx.isdigit() and 1 <= int(idx) <= len(sheet_names):
                return sheet_names[int(idx) - 1]
            print("序号无效，请重新输入。")
            continue

        print("请输入 1 或 2。")


def convert_xlsx_to_csv(xlsx_path: Path, output_dir: Path, selected_sheet: Optional[str] = None) -> List[Path]:
    wb = load_workbook(xlsx_path, data_only=True)
    created_files: List[Path] = []

    try:
        sheet_names = [selected_sheet] if selected_sheet else wb.sheetnames
        multi_output = len(sheet_names) > 1

        for sheet_name in sheet_names:
            ws = wb[sheet_name]

            if multi_output:
                file_name = f"{xlsx_path.stem}_{safe_name(sheet_name)}.csv"
            else:
                file_name = f"{xlsx_path.stem}.csv"

            out_path = output_dir / file_name
            with out_path.open("w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if v is None else v for v in row])

            created_files.append(out_path)
    finally:
        wb.close()

    return created_files


def main() -> None:
    print("=== XLSX 转 CSV 工具 ===")

    xlsx_path = ask_xlsx_path()
    output_dir = ask_output_dir(xlsx_path.parent)

    wb = load_workbook(xlsx_path, read_only=True)
    try:
        selected_sheet = ask_sheet_mode(wb.sheetnames)
    finally:
        wb.close()

    files = convert_xlsx_to_csv(xlsx_path, output_dir, selected_sheet)

    print("\n转换完成，已生成以下文件：")
    for f in files:
        print(f"- {f}")


if __name__ == "__main__":
    main()
